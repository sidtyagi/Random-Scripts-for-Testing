'''
Created on 13-Mar-2017

@author: sidtyagi
'''
'''
Created on 13-Mar-2017

@author: sidtyagi
'''
'''
Created on 13-Mar-2017

@author: sidtyagi
'''
'''
Created on 13-Mar-2017

@author: sidtyagi
'''
import xml.etree.ElementTree as ET
import xlrd,openpyxl,time
from openpyxl.styles import Alignment
import os


def write_to_excel(row,col,config,sheet_name):
    print 'writing to row '+str(row)
    print 'writing to col '+str(col)
    #----------------------------#
    xfile = openpyxl.load_workbook('TEST.xlsx') 
    
    if sheet_name not in xfile.sheetnames:
        new_sheet = xfile.create_sheet(sheet_name)
        new_sheet.cell(row=row, column=col).value = config    
        xfile.save('TEST.xlsx')
    
    else:        
        #new_sheet = xfile.get_sheet_by_name('Sheet1')
        new_sheet = xfile.get_sheet_by_name(sheet_name)    
        #-----------------------------#  
        #row_no = sheet.max_row + 1
        #column_no = sheet.max_column
        new_sheet.cell(row=row, column=col).value = config    
        xfile.save('TEST.xlsx') 



def fetch_transit_hops(stuff):
    path_only = stuff.findall('configuration/protocols/mpls/path')
    #--------------get all paths with hops as a dict and not the lsp path----#
    print path_only
    dict_path={}
    for p1 in path_only:
        path_hops=[]
        #p1 represents one path which will have path name, path list
        p2=p1._children
        #now using p2 we r traversing the path childern
        path_name=''
        path_name=p2[0].text
        print path_name
        #first entity is the path name
        
        #----now using below fetching the hops and strict/loose
        for p3 in p2:
            p4=p3._children
            for p5 in p4:            
                if 'strict' in p5.tag or 'loose' in p5.tag:
                    path_hops.append(p5.tag)            
                if p5.text is not None:  
                    if 'strict' in p5.text or 'loose' in p5.text or '.' in p5.text:                                    
                        path_hops.append(p5.text)
                        
        dict_path[path_name]=path_hops
    print dict_path
    return dict_path
        

#XML DATA folder should only contain the text files
all_text_files=os.listdir('XML DATA')
print all_text_files

#f=open('XML DATA//BPL-MPL-PE-RTR-37-32_10.123.37.32.txt')
for f_name in all_text_files:
    print f_name
    print '&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&'
    
    f=open('XML DATA//'+f_name)
    f_content=''
    f_content=f.read()
    #print f_content
    try:
        stuff=''
        stuff = ET.fromstring(f_content)
        dict_path=''
        dict_path=fetch_transit_hops(stuff)
        
        label_switch_path = stuff.findall('configuration/protocols/mpls/label-switched-path')
        row=0
        col=1
        for x in label_switch_path:
            row+=1  
            col=1    
            y= x._children
            print y    
            for item in y:              
                temp_str=str(item)
                if "Element 'name'" in temp_str:
                    print temp_str
                    label_switch_path_name=item.text
                    print 'writing '+label_switch_path_name
                    write_to_excel(row, col, label_switch_path_name,f_name.replace('.txt',''))         
                if "Element 'from'" in temp_str:
                    lsp_source=item.text 
                    col+=1
                    write_to_excel(row, col, lsp_source,f_name.replace('.txt','')) 
                         
                if "Element 'to'" in temp_str:
                    lsp_destination= item.text
                    col+=1           
                    write_to_excel(row, col, lsp_destination,f_name.replace('.txt',''))
                
                
                if "Element 'primary'" in temp_str:
                    k1=item._children
                    for k2 in k1:
                        temp_str2=str(k2)
                        if "Element 'name'" in temp_str2:
                            pri_path_name=k2.text                                      
                            col=4
                            row+=1
                            write_to_excel(row, col, pri_path_name,f_name.replace('.txt',''))
                            
                            #---write the state of this path to excel----#
                            col=5                    
                            write_to_excel(row, col, "primary",f_name.replace('.txt',''))
                            
                            #-------write the transit hops to excel----#
                            transit_hops=''
                            transit_hops=dict_path[pri_path_name]
                            
                            for hop in transit_hops:
                                col+=1
                                write_to_excel(row, col, hop,f_name.replace('.txt',''))
                        
                                
                
                if "Element 'secondary'" in temp_str:
                    k3=item._children
                    for k4 in k3:
                        temp_str3=str(k4)                
                        
                        if "Element 'name'" in temp_str3:
                            sec_path_name=k4.text                    
                            col=4
                            row+=1
                            write_to_excel(row, col, sec_path_name,f_name.replace('.txt',''))                        
                            #-----Now write the transit hops to excel------#
                            transit_hops_sec=''
                            transit_hops_sec=dict_path[sec_path_name]
                            col=5
                            for hop in transit_hops_sec:
                                #----start writing from the 6th column to excel sheet to push transit rows--#
                                col+=1
                                write_to_excel(row, col, hop,f_name.replace('.txt',''))
                                
                        if "Element 'standby'" in temp_str3:                                        
                            col=5                    
                            write_to_excel(row, col, 'standby',f_name.replace('.txt',''))
        
    except:
        f_fail=open('FAILED FOR.txt','a')
        f_fail.write(f_name)   
        f_fail.close()     
